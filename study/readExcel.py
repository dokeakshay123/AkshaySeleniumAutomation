import openpyxl

file="D://test//data.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row # 4 count of row

column=sheet.max_column # 5 count of column

# reading all the rows and column from excel sheet
for r in range(1,rows+1):
    for c in range(1,column+1):
        print(sheet.cell(r,c).value,end='      ')

    print()


