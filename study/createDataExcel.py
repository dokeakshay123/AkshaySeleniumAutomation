import openpyxl

# file="D:/test/createData.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook.active  # or sheet=workbook["Sheet1"] -- get active sheet from excel
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Akshay"
#         print(r,c)
#
# workbook.save(file)

#  multiple data
file="D:/test/createData.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active  # or sheet=workbook["Sheet1"] -- get active sheet from excel


sheet.cell(1,1).value="Akshay"
sheet.cell(1,2).value="D"
sheet.cell(1,3).value=123

sheet.cell(2,1).value="Teja"
sheet.cell(2,2).value="D"
sheet.cell(2,3).value=5678

workbook.save(file)